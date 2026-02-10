# core/excel_reader.py

from openpyxl import load_workbook

class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None

    def load(self):
        self.workbook = load_workbook(
            filename=self.file_path,
            data_only=False
        )
        return self.workbook

    def get_sheets(self):
        if not self.workbook:
            raise RuntimeError("Workbook not loaded")
        return self.workbook.sheetnames

    def get_sheet(self, sheet_name):
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        return self.workbook[sheet_name]
