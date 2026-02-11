from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook


class BotState(str, Enum):
    WAIT_FILE = "WAIT_FILE"
    ANALYZED = "ANALYZED"
    SELECT_OPERATION = "SELECT_OPERATION"
    SELECT_ROW = "SELECT_ROW"
    SELECT_COLUMN = "SELECT_COLUMN"
    INPUT_TEXT = "INPUT_TEXT"
    CONFIRM_OPERATION = "CONFIRM_OPERATION"
    READY_TO_SAVE = "READY_TO_SAVE"


@dataclass
class PendingOperation:
    op_kind: str  # add / delete / edit
    target_kind: str  # row / column
    mode: str  # single / group
    selected: set[int] = field(default_factory=set)
    payload_lines: list[str] = field(default_factory=list)


@dataclass
class SessionData:
    state: BotState = BotState.WAIT_FILE
    ui_mode: str = "inline"
    original_file_name: str | None = None
    original_bytes: bytes | None = None
    working_bytes: bytes | None = None
    selected_sheet: str | None = None
    op_stack: list[PendingOperation] = field(default_factory=list)
    undo_stack: list[bytes] = field(default_factory=list)
    pending: PendingOperation | None = None


class SessionManager:
    def __init__(self):
        self._store: dict[int, SessionData] = {}

    def get(self, chat_id: int) -> SessionData:
        return self._store.setdefault(chat_id, SessionData())

    def clear_after_save(self, chat_id: int):
        session = self.get(chat_id)
        session.op_stack.clear()
        session.undo_stack.clear()
        session.pending = None
        session.state = BotState.WAIT_FILE


def _wb_from_bytes(data: bytes):
    return load_workbook(BytesIO(data))


def _bytes_from_wb(wb) -> bytes:
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def analyze_workbook(file_bytes: bytes) -> dict:
    wb = _wb_from_bytes(file_bytes)
    result = {"sheets": []}
    for sheet in wb.worksheets:
        row_count = sheet.max_row
        col_count = sheet.max_column
        headers = [sheet.cell(row=1, column=i).value for i in range(1, col_count + 1)]
        result["sheets"].append(
            {
                "name": sheet.title,
                "rows": row_count,
                "cols": col_count,
                "headers": headers,
            }
        )
    return result


def get_sheet_map(file_bytes: bytes, sheet_name: str) -> tuple[list[str], list[int]]:
    wb = _wb_from_bytes(file_bytes)
    sheet = wb[sheet_name]
    headers = [str(sheet.cell(row=1, column=i).value or f"Column_{i}") for i in range(1, sheet.max_column + 1)]
    rows = list(range(2, sheet.max_row + 1))
    return headers, rows


def apply_operation(working_bytes: bytes, sheet_name: str, op: PendingOperation) -> bytes:
    wb = _wb_from_bytes(working_bytes)
    sheet = wb[sheet_name]

    if not op.selected:
        raise ValueError("هیچ آیتمی انتخاب نشده است")

    selected = sorted(op.selected)

    if op.target_kind == "column":
        if any(i < 1 or i > sheet.max_column for i in selected):
            raise ValueError("ستون انتخاب‌شده معتبر نیست (احتمالاً قبلاً حذف شده)")
    else:
        if any(i < 2 or i > sheet.max_row for i in selected):
            raise ValueError("سطر انتخاب‌شده معتبر نیست (احتمالاً قبلاً حذف شده)")

    if op.op_kind == "delete":
        if op.target_kind == "column":
            for idx in sorted(selected, reverse=True):
                sheet.delete_cols(idx, 1)
        else:
            for idx in sorted(selected, reverse=True):
                sheet.delete_rows(idx, 1)
        return _bytes_from_wb(wb)

    if op.op_kind in {"add", "edit"}:
        if not op.payload_lines:
            raise ValueError("متنی برای عملیات ثبت نشده است")

        if op.op_kind == "edit":
            text = "\n".join(op.payload_lines).strip()
            if op.target_kind == "column":
                for col in selected:
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(row=row, column=col).value = text
            else:
                for row_idx in selected:
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row_idx, column=col).value = text
            return _bytes_from_wb(wb)

        # add
        if op.target_kind == "column":
            if op.mode == "single":
                text = "\n".join(op.payload_lines).strip()
                for col in selected:
                    target_row = sheet.max_row + 1
                    sheet.cell(row=target_row, column=col).value = text
            else:
                lines = [line for line in op.payload_lines if line.strip()]
                for col in selected:
                    for line in lines:
                        target_row = sheet.max_row + 1
                        sheet.cell(row=target_row, column=col).value = line
            return _bytes_from_wb(wb)

        # row target
        if op.mode == "single":
            text = "\n".join(op.payload_lines).strip()
            for row_idx in selected:
                target_col = sheet.max_column + 1
                sheet.cell(row=row_idx, column=target_col).value = text
        else:
            lines = [line for line in op.payload_lines if line.strip()]
            for row_idx in selected:
                col = sheet.max_column + 1
                for line in lines:
                    sheet.cell(row=row_idx, column=col).value = line
                    col += 1
        return _bytes_from_wb(wb)

    raise ValueError("نوع عملیات پشتیبانی نمی‌شود")


def save_working_copy(working_bytes: bytes, original_name: str) -> Path:
    suffix = Path(original_name).suffix or ".xlsx"
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(working_bytes)
        return Path(tmp.name)
