from __future__ import annotations

import hashlib
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook


class ExcelEngine:
    def __init__(self, base_dir: str = "storage/uploads"):
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)

    def store_original_and_working(self, user_id: int, filename: str, content: bytes) -> tuple[str, str, str]:
        file_hash = hashlib.sha256(content).hexdigest()[:16]
        file_id = f"u{user_id}_{file_hash}"
        original_path = self.base_dir / f"{file_id}_original_{filename}"
        working_path = self.base_dir / f"{file_id}_working_{filename}"
        original_path.write_bytes(content)
        working_path.write_bytes(content)
        return file_id, str(original_path), str(working_path)

    def analyze(self, file_path: str) -> dict:
        wb = load_workbook(file_path)
        sheets = []
        for ws in wb.worksheets:
            headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
            sheets.append(
                {
                    "name": ws.title,
                    "rows": ws.max_row,
                    "cols": ws.max_column,
                    "headers": headers,
                }
            )
        return {"sheets": sheets}

    def filter_contains(self, working_path: str, sheet_name: str, column_name: str, keyword: str):
        wb = load_workbook(working_path)
        ws = wb[sheet_name]
        headers = [str(ws.cell(1, i).value or "") for i in range(1, ws.max_column + 1)]
        if column_name not in headers:
            raise ValueError(f"ستون {column_name} یافت نشد")
        col_idx = headers.index(column_name) + 1

        delete_rows = []
        for r in range(2, ws.max_row + 1):
            val = str(ws.cell(r, col_idx).value or "")
            if keyword not in val:
                delete_rows.append(r)

        for r in reversed(delete_rows):
            ws.delete_rows(r, 1)

        wb.save(working_path)

    def export(self, working_path: str, output_name: str) -> str:
        out_path = self.base_dir / output_name
        shutil.copy2(working_path, out_path)
        return str(out_path)

    @staticmethod
    def dumps_analysis(analysis: dict) -> str:
        return json.dumps(analysis, ensure_ascii=False)
