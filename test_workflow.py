from io import BytesIO

from openpyxl import Workbook, load_workbook

from bot.workflow import PendingOperation, analyze_workbook, apply_operation


def build_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["name", "price"])
    ws.append(["a", 10])
    ws.append(["b", 20])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_analyze_workbook():
    data = build_bytes()
    info = analyze_workbook(data)
    assert len(info["sheets"]) == 1
    assert info["sheets"][0]["rows"] == 3


def test_apply_add_group_column():
    data = build_bytes()
    op = PendingOperation(op_kind="add", target_kind="column", mode="group", selected={2}, payload_lines=["x", "y", ""])
    out = apply_operation(data, "Sheet", op)
    wb = load_workbook(BytesIO(out))
    ws = wb["Sheet"]
    assert ws.cell(row=4, column=2).value == "x"
    assert ws.cell(row=5, column=2).value == "y"


def test_apply_delete_row():
    data = build_bytes()
    op = PendingOperation(op_kind="delete", target_kind="row", mode="single", selected={2})
    out = apply_operation(data, "Sheet", op)
    wb = load_workbook(BytesIO(out))
    ws = wb["Sheet"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "b"
